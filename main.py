import os
import sys
import html
import sqlite3
import math
import calendar
import hmac
import hashlib
import logging
import asyncio
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from dotenv import load_dotenv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from aiogram import Bot, Dispatcher, F, types
from aiogram.filters import CommandStart, Command
from aiogram.types import (
    ReplyKeyboardMarkup,
    KeyboardButton,
    FSInputFile,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    WebAppInfo,
)
from aiogram.exceptions import TelegramForbiddenError
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

from apscheduler.schedulers.asyncio import AsyncIOScheduler

# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s"
)
logger = logging.getLogger("attendance_bot")

# ================= CONFIGURATION & ENV =================
load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS")

if not BOT_TOKEN:
    sys.exit(
        "XATOLIK: BOT_TOKEN topilmadi!\n"
        "Iltimos, loyihaning ildizida .env fayl yarating va unga quyidagini yozing:\n"
        "BOT_TOKEN=your_telegram_bot_token\n"
        "ADMIN_IDS=123456789,987654321"
    )

if not ADMIN_IDS_RAW:
    sys.exit(
        "XATOLIK: ADMIN_IDS topilmadi!\n"
        "Iltimos, .env fayliga admin(lar) Telegram ID sini kiriting:\n"
        "ADMIN_IDS=123456789,987654321"
    )

try:
    ADMIN_IDS = [int(x.strip()) for x in ADMIN_IDS_RAW.split(",") if x.strip()]
except ValueError:
    sys.exit("XATOLIK: ADMIN_IDS faqat vergul bilan ajratilgan raqamlardan iborat bo'lishi kerak.")

if not ADMIN_IDS:
    sys.exit("XATOLIK: ADMIN_IDS bo'sh bo'lishi mumkin emas.")

OFFICE_LAT = float(os.getenv("OFFICE_LAT", "41.3199585"))
OFFICE_LON = float(os.getenv("OFFICE_LON", "69.2661517"))
ALLOWED_RADIUS_METERS = int(os.getenv("ALLOWED_RADIUS_METERS", "100"))
WORK_START_TIME = os.getenv("WORK_START_TIME", "09:00")
WORK_END_TIME = os.getenv("WORK_END_TIME", "18:00")
# Admin uchun mini web-ilova (Telegram WebApp) manzili. Bo'sh bo'lsa,
# admin klaviaturasida "Web Panel" tugmasi ko'rsatilmaydi.
WEBAPP_URL = os.getenv("WEBAPP_URL", "").strip()
TASHKENT_TZ = ZoneInfo("Asia/Tashkent")


def _webapp_sig(admin_id: int) -> str:
    """Telegram initData ba'zi klientlarda (masalan Desktop) bo'sh kelishi
    mumkinligi aniqlangani uchun, Web Panel kirishini Telegram'ga bog'liq
    bo'lmagan holda o'zimiz imzolaymiz (BOT_TOKEN sir kaliti sifatida)."""
    return hmac.new(BOT_TOKEN.encode(), str(admin_id).encode(), hashlib.sha256).hexdigest()


def _webapp_url_for(admin_id: int) -> str:
    """Diqqat: Telegram (ayniqsa Desktop) tugma manzilidagi "?..." so'rov
    qismini butunlay olib tashlashi kuzatilgan, shuning uchun uid/sig'ni
    so'rov parametri sifatida EMAS, manzil YO'LI (path) ichiga qo'yamiz -
    bu qism Telegram tomonidan o'zgartirilmaydi."""
    if not WEBAPP_URL:
        return ""
    base = WEBAPP_URL.rstrip("/")
    return f"{base}/a/{admin_id}/{_webapp_sig(admin_id)}"

DB_PATH = os.getenv("DB_PATH", "attendance.db")


def esc(value) -> str:
    """HTML uchun xavfsiz matn (foydalanuvchi kiritgan matnlarni escape qiladi)."""
    return html.escape(str(value)) if value is not None else ""


# ================= FSM STATES =================
class AttendanceState(StatesGroup):
    waiting_for_reason = State()
    waiting_for_checkout_location = State()
    waiting_for_broadcast_text = State()


# ================= DATABASE SETUP & HELPERS =================
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_now():
    return datetime.now(TASHKENT_TZ)


async def run_db(func, *args, **kwargs):
    """Bloklovchi SQLite chaqiruvlarini alohida thread'da bajaradi (event loop'ni bloklamaslik uchun)."""
    return await asyncio.to_thread(func, *args, **kwargs)


def init_db():
    with get_db() as conn:
        cursor = conn.cursor()

        # Users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                full_name TEXT,
                is_approved INTEGER DEFAULT 0,
                monthly_salary REAL DEFAULT 0.0,
                norm_days INTEGER DEFAULT 26,
                work_start_time TEXT DEFAULT NULL,
                work_end_time TEXT DEFAULT NULL
            )
        ''')

        # Attendance table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                date TEXT,
                check_in_time TEXT,
                check_out_time TEXT,
                lateness_minutes INTEGER DEFAULT 0,
                lateness_reason TEXT DEFAULT '',
                break_start TEXT DEFAULT NULL,
                break_end TEXT DEFAULT NULL,
                break_minutes INTEGER DEFAULT 0,
                work_hours REAL DEFAULT 0,
                checkin_source TEXT DEFAULT 'self',
                checkout_source TEXT DEFAULT 'self',
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        # Advances table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                amount REAL DEFAULT 0.0,
                date TEXT,
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        # Ish jadvali (kalendar): xodim uchun BELGILANGAN ish kunlari.
        # Bir qator = shu xodim uchun shu sana "ish kuni" deb belgilangan.
        # Biror oy uchun bu yerda yozuv bo'lmasa, users.norm_days standart
        # qiymat sifatida ishlatiladi (orqaga moslik uchun).
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS work_schedule (
                user_id INTEGER,
                date TEXT,
                PRIMARY KEY (user_id, date),
                FOREIGN KEY(user_id) REFERENCES users(user_id)
            )
        ''')

        # Eski bazalarni xavfsiz yangilash (Migration)
        columns_to_add = [
            ("users", "is_approved INTEGER DEFAULT 0"),
            ("users", "monthly_salary REAL DEFAULT 0.0"),
            ("users", "norm_days INTEGER DEFAULT 26"),
            ("users", "work_start_time TEXT DEFAULT NULL"),
            ("users", "work_end_time TEXT DEFAULT NULL"),
            ("attendance", "lateness_reason TEXT DEFAULT ''"),
            ("attendance", "break_start TEXT DEFAULT NULL"),
            ("attendance", "break_end TEXT DEFAULT NULL"),
            ("attendance", "break_minutes INTEGER DEFAULT 0"),
            ("attendance", "checkin_source TEXT DEFAULT 'self'"),
            ("attendance", "checkout_source TEXT DEFAULT 'self'")
        ]

        for table, col in columns_to_add:
            try:
                cursor.execute(f"ALTER TABLE {table} ADD COLUMN {col}")
            except sqlite3.OperationalError:
                pass

        # Bir kunda bitta xodim uchun faqat bitta davomat yozuvi bo'lishini
        # ta'minlaydigan UNIQUE indeks (race condition'ni oldini oladi).
        # Eslatma: agar bazada avvaldan dublikatlar bo'lsa, bu buyruq xato beradi -
        # bu holda avval dublikatlarni qo'lda tozalash kerak bo'ladi.
        try:
            cursor.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_attendance_user_date "
                "ON attendance(user_id, date)"
            )
        except sqlite3.IntegrityError:
            logger.warning(
                "attendance jadvalida (user_id, date) bo'yicha dublikat yozuvlar topildi. "
                "UNIQUE indeks yaratilmadi - avval dublikatlarni tozalang."
            )

        for admin_id in ADMIN_IDS:
            cursor.execute(
                "INSERT OR IGNORE INTO users (user_id, full_name, is_approved, monthly_salary, norm_days) "
                "VALUES (?, ?, 1, 0.0, 26)",
                (admin_id, "Admin")
            )
            # Har doim admin huquqini tasdiqlangan holatga o'tkazamiz
            cursor.execute("UPDATE users SET is_approved = 1 WHERE user_id = ?", (admin_id,))
        conn.commit()


def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371000  # Yer radiusi metrlarda
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)

    a = math.sin(delta_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


# ================= ISH KUNLARI KALENDARI (NORMA) =================
def _norm_days_for_month(cursor, user_id: int, month_prefix: str, fallback_norm_days) -> int:
    """Shu xodim uchun shu oyda (Web Panel kalendarida) belgilangan ish
    kunlari sonini qaytaradi. Agar hech qanday kun belgilanmagan bo'lsa,
    users.norm_days (yoki 26) standart qiymat sifatida ishlatiladi -
    shunday qilib eski xodimlar uchun hech narsa buzilmaydi."""
    cursor.execute(
        "SELECT COUNT(*) FROM work_schedule WHERE user_id = ? AND date LIKE ?",
        (user_id, f"{month_prefix}%")
    )
    count = cursor.fetchone()[0]
    if count and count > 0:
        return count
    return fallback_norm_days if fallback_norm_days and fallback_norm_days > 0 else 26


def _get_work_days_sync(user_id: int, month_prefix: str):
    """Shu xodim uchun shu oyda belgilangan barcha ish kunlari sanalarini qaytaradi."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT date FROM work_schedule WHERE user_id = ? AND date LIKE ? ORDER BY date",
            (user_id, f"{month_prefix}%")
        )
        return [r[0] for r in cursor.fetchall()]


def _toggle_work_day_sync(user_id: int, date_str: str) -> bool:
    """Berilgan sanani ish kuni deb belgilaydi (agar belgilanmagan bo'lsa) yoki
    belgisini olib tashlaydi (agar allaqachon belgilangan bo'lsa).
    Qaytadi: True - endi ish kuni deb belgilandi, False - belgi olib tashlandi."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT 1 FROM work_schedule WHERE user_id = ? AND date = ?", (user_id, date_str)
        )
        exists = cursor.fetchone()
        if exists:
            cursor.execute(
                "DELETE FROM work_schedule WHERE user_id = ? AND date = ?", (user_id, date_str)
            )
            conn.commit()
            return False
        else:
            cursor.execute(
                "INSERT INTO work_schedule (user_id, date) VALUES (?, ?)", (user_id, date_str)
            )
            conn.commit()
            return True


def _set_work_days_sync(user_id: int, month_prefix: str, days: list[int]):
    """Shu oy uchun xodimning ish kunlarini TO'LIQ shu ro'yxat bilan almashtiradi
    (avvalgi belgilar shu oy uchun o'chirilib, yangilari yoziladi)."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "DELETE FROM work_schedule WHERE user_id = ? AND date LIKE ?",
            (user_id, f"{month_prefix}%")
        )
        for day in days:
            date_str = f"{month_prefix}-{int(day):02d}"
            cursor.execute(
                "INSERT OR IGNORE INTO work_schedule (user_id, date) VALUES (?, ?)",
                (user_id, date_str)
            )
        conn.commit()


init_db()

# ================= BOT & SCHEDULER INITIALIZATION =================
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# DIQQAT: AsyncIOScheduler'ni bu yerda (modul import qilinayotganda) YARATISH
# XATOLIKKA OLIB KELADI: u shu paytda "joriy" asyncio event loop'ni ushlab
# oladi (masalan Flask/Passenger so'rov ishlaydigan asosiy thread'nikini),
# bot esa keyinchalik BUTUNLAY BOSHQA thread'da, YANGI event loop bilan
# ishga tushadi (asyncio.run(main())). Natijada scheduler.add_job() chaqirilganda
# "RuntimeError: Event loop is closed" xatosi chiqadi. Shu sabab scheduler
# pastda main() ichida, bot o'zining haqiqiy event loop'ini olgandan KEYIN
# yaratiladi.
scheduler = None

# ARXITEKTURA ESLATMASI: bot (main.py) va admin Web Panel (webapp.py) endi
# IKKITA MUSTAQIL jarayon sifatida ishlaydi (bot - cron+nohup orqali,
# webapp - Passenger orqali) - shuning uchun ular bir xil Python
# jarayoni/thread'ida EMAS. Shu sabab Web Panel botning xabar yuborish
# funksiyasini (bot.send_message) to'g'ridan-to'g'ri chaqira olmaydi -
# buning o'rniga webapp.py Telegram Bot API'ga oddiy HTTP so'rov yuboradi
# (bot jarayonidan mustaqil holda). Shuning uchun BOT_LOOP/asyncio ko'prigi
# endi kerak emas.

# ================= KEYBOARDS =================
_user_rows = [
    [KeyboardButton(text="🟢 Ishga keldim"), KeyboardButton(text="🔴 Ishdan ketdim")],
    [KeyboardButton(text="☕ Tanaffusga chiqdim"), KeyboardButton(text="🏢 Tanaffusdan qaytdim")],
    [KeyboardButton(text="📊 Mening statistikam"), KeyboardButton(text="💰 Mening oyligim")],
]

user_keyboard = ReplyKeyboardMarkup(keyboard=_user_rows, resize_keyboard=True)

def _base_admin_rows():
    return [
        [KeyboardButton(text="🟢 Ishga keldim"), KeyboardButton(text="🔴 Ishdan ketdim")],
        [KeyboardButton(text="☕ Tanaffusga chiqdim"), KeyboardButton(text="🏢 Tanaffusdan qaytdim")],
        [KeyboardButton(text="📊 Mening statistikam"), KeyboardButton(text="💰 Mening oyligim")],
        [KeyboardButton(text="👥 Xodimlar"), KeyboardButton(text="📈 Barcha davomat hisoboti (Excel)")],
        [KeyboardButton(text="📢 Xabarnoma yuborish")],
    ]


# Web Panel tugmasidagi manzil har bir admin uchun ALOHIDA (o'zining imzosi
# bilan) - shuning uchun har bir admin uchun alohida klaviatura tayyorlanadi.
_admin_keyboards = {}
for _admin_id in ADMIN_IDS:
    _rows = _base_admin_rows()
    if WEBAPP_URL:
        _rows.append(
            [KeyboardButton(text="🖥 Web Panel", web_app=WebAppInfo(url=_webapp_url_for(_admin_id)))]
        )
    _admin_keyboards[_admin_id] = ReplyKeyboardMarkup(keyboard=_rows, resize_keyboard=True)

# Orqaga qarab moslik uchun: eski kodda ishlatiladigan yagona `admin_keyboard`
# (masalan hali /start bosmagan admin uchun) - Web Panel tugmasisiz.
admin_keyboard = ReplyKeyboardMarkup(keyboard=_base_admin_rows(), resize_keyboard=True)

location_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📍 Joylashuvni yuborish", request_location=True)]
    ],
    resize_keyboard=True,
    one_time_keyboard=True
)


def get_kb_for(user_id: int):
    if user_id in ADMIN_IDS:
        return _admin_keyboards.get(user_id, admin_keyboard)
    return user_keyboard


def _is_approved_sync(user_id: int) -> bool:
    """Foydalanuvchi ro'yxatdan o'tgan va admin tomonidan tasdiqlangan-tasdiqlanmaganini tekshiradi."""
    if user_id in ADMIN_IDS:
        return True
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT is_approved FROM users WHERE user_id = ?", (user_id,))
        row = cursor.fetchone()
        return bool(row and row[0] == 1)


async def is_approved(user_id: int) -> bool:
    return await run_db(_is_approved_sync, user_id)


# ================= EXCEL GENERATOR =================
def _generate_excel_report_sync(month_prefix: str, file_path: str):
    year, month = (int(x) for x in month_prefix.split("-"))
    days_in_month = calendar.monthrange(year, month)[1]

    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT u.user_id, u.full_name, u.monthly_salary, u.norm_days
            FROM users u WHERE u.is_approved = 1
            ORDER BY u.full_name COLLATE NOCASE
        ''')
        users = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Oylik Hisobot"

        headers = [
            "Xodim F.I.Sh", "Belgilangan Oylik", "Norma kun",
            "Jami Ishlangan Soat", "Hisoblangan Maosh", "Berilgan Avans", "Sof Beriladigan Oylik"
        ]
        for day in range(1, days_in_month + 1):
            headers.append(f"{day}-kun Kelgan")
            headers.append(f"{day}-kun Ketgan")
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for u_id, full_name, m_salary, fallback_norm_d in users:
            cursor.execute(
                "SELECT COALESCE(SUM(work_hours), 0.0) FROM attendance WHERE user_id = ? AND date LIKE ?",
                (u_id, f"{month_prefix}%")
            )
            w_hours = cursor.fetchone()[0] or 0.0

            cursor.execute(
                "SELECT COALESCE(SUM(amount), 0.0) FROM advances WHERE user_id = ? AND date LIKE ?",
                (u_id, f"{month_prefix}%")
            )
            adv_sum = cursor.fetchone()[0] or 0.0

            norm_d = _norm_days_for_month(cursor, u_id, month_prefix, fallback_norm_d)
            hourly_rate = m_salary / (norm_d * 8) if norm_d and norm_d > 0 else 0
            earned = w_hours * hourly_rate
            net_salary = earned - adv_sum

            row = [
                full_name, m_salary, norm_d,
                round(w_hours, 1), round(earned, 2), adv_sum, round(net_salary, 2)
            ]

            # Shu oy uchun xodimning har bir kunlik kelish/ketish vaqtlarini
            # (date -> (check_in, check_out)) xaritaga yig'ib olamiz.
            cursor.execute(
                "SELECT date, check_in_time, check_out_time FROM attendance "
                "WHERE user_id = ? AND date LIKE ?",
                (u_id, f"{month_prefix}%")
            )
            daily_map = {d: (ci, co) for d, ci, co in cursor.fetchall()}

            for day in range(1, days_in_month + 1):
                date_str = f"{month_prefix}-{day:02d}"
                check_in, check_out = daily_map.get(date_str, (None, None))
                row.append(check_in[:5] if check_in else "-")
                row.append(check_out[:5] if check_out else "-")

            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        wb.save(file_path)


async def generate_excel_report(month_prefix: str, file_path: str):
    await run_db(_generate_excel_report_sync, month_prefix, file_path)


# ---- Kunlik (bir kunlik) davomat hisoboti - Excel ----
def _generate_daily_excel_report_sync(date_str: str, file_path: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT u.user_id, u.full_name, a.check_in_time, a.check_out_time, a.lateness_minutes,
                   a.lateness_reason, a.break_minutes, a.work_hours, u.monthly_salary, u.norm_days
            FROM attendance a
            JOIN users u ON u.user_id = a.user_id
            WHERE a.date = ?
            ORDER BY u.full_name COLLATE NOCASE
        ''', (date_str,))
        rows = cursor.fetchall()

        # Bugun umuman check-in qilmagan tasdiqlangan xodimlarni ham aniqlaymiz
        cursor.execute('''
            SELECT u.full_name FROM users u
            WHERE u.is_approved = 1
              AND u.user_id NOT IN (SELECT user_id FROM attendance WHERE date = ?)
            ORDER BY u.full_name COLLATE NOCASE
        ''', (date_str,))
        absent = [r[0] for r in cursor.fetchall()]

        # Har bir xodim uchun shu oyning haqiqiy ish kuni normasini (bog'lanish
        # yopilmasdan oldin) hisoblab olamiz.
        month_prefix = date_str[:7]
        norm_map = {
            u_id: _norm_days_for_month(cursor, u_id, month_prefix, fallback_norm_d)
            for u_id, _fn, _ci, _co, _lt, _rs, _bm, _wh, _ms, fallback_norm_d in rows
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kunlik Hisobot"

    headers = [
        "Xodim F.I.Sh", "Kelgan vaqti", "Ketgan vaqti", "Kechikish (daq)",
        "Kechikish sababi", "Tanaffus (daq)", "Ishlangan soat", "Kunlik topilgan pul"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for u_id, full_name, check_in, check_out, lateness, reason, break_min, work_hours, m_salary, _fallback in rows:
        norm_d = norm_map.get(u_id, 26)
        hourly_rate = (m_salary / (norm_d * 8)) if norm_d and norm_d > 0 else 0
        daily_earned = (work_hours or 0) * hourly_rate
        ws.append([
            full_name,
            check_in or "-",
            check_out or "Hali ketmagan",
            lateness or 0,
            reason or "",
            break_min or 0,
            round(work_hours or 0, 2),
            round(daily_earned, 2)
        ])

    if absent:
        absent_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
        for name in absent:
            row_idx = ws.max_row + 1
            ws.append([name, "KELMADI", "-", 0, "", 0, 0, 0])
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_num).fill = absent_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)


async def generate_daily_excel_report(date_str: str, file_path: str):
    await run_db(_generate_daily_excel_report_sync, date_str, file_path)


# ================= AUTOMATIC JOBS (SCHEDULER) =================
def _get_checkin_reminder_targets_sync(today_str: str, now_hm: str):
    """Har bir xodimning SHAXSIY ish boshlash vaqtiga nisbatan 15, 10 va 0 daqiqa
    qolganda (hali 'Ishga keldim' bosilmagan bo'lsa) eslatma yuborish kerak
    bo'lgan xodimlarni topadi."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, full_name, work_start_time FROM users WHERE is_approved = 1")
        users = cursor.fetchall()
        targets = []
        for user_id, full_name, work_start in users:
            start = work_start or WORK_START_TIME
            try:
                start_dt = datetime.strptime(start, "%H:%M")
            except ValueError:
                continue
            for offset in (15, 10, 0):
                target_hm = (start_dt - timedelta(minutes=offset)).strftime("%H:%M")
                if target_hm == now_hm:
                    cursor.execute(
                        "SELECT id FROM attendance WHERE user_id = ? AND date = ?",
                        (user_id, today_str)
                    )
                    if not cursor.fetchone():
                        targets.append((user_id, full_name, offset))
                    break
        return targets


async def send_checkin_reminders():
    """Har daqiqada ishga tushadi: har bir xodimning SHAXSIY ish boshlash
    vaqtiga nisbatan 15, 10 va 0 daqiqa qolganda (masalan 9:00 dan ishlasa -
    8:45, 8:50, 9:00 da) hali kelmagan bo'lsa eslatma yuboradi."""
    now = get_now()
    today_str = now.strftime("%Y-%m-%d")
    now_hm = now.strftime("%H:%M")
    targets = await run_db(_get_checkin_reminder_targets_sync, today_str, now_hm)

    for user_id, full_name, offset in targets:
        try:
            kb = get_kb_for(user_id)
            if offset > 0:
                body = f"Ish vaqti boshlanishiga <b>{offset} daqiqa</b> qoldi."
            else:
                body = "Ish vaqtingiz boshlandi."
            await bot.send_message(
                chat_id=user_id,
                text=(
                    f"Xayrli kun, <b>{esc(full_name)}</b>! ☀️\n"
                    f"{body}\n"
                    f"Ofisga kelgach, <b>🟢 Ishga keldim</b> tugmasini bosishni unutmang!"
                ),
                reply_markup=kb,
                parse_mode="HTML"
            )
        except TelegramForbiddenError:
            logger.info(f"Foydalanuvchi {user_id} botni bloklagan, eslatma yuborilmadi.")
        except Exception as e:
            logger.error(f"Eslatma yuborishda xatolik ({user_id}): {e}")


def _get_checkout_reminder_targets_sync(today_str: str, now_hm: str):
    """Har bir xodimning SHAXSIY ish tugash vaqtiga nisbatan 15, 10 va 5 daqiqa
    qolganda (bugun kelgan, lekin hali 'Ishdan ketdim' bosilmagan bo'lsa)
    eslatma yuborish kerak bo'lgan xodimlarni topadi."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, full_name, work_end_time FROM users WHERE is_approved = 1")
        users = cursor.fetchall()
        targets = []
        for user_id, full_name, work_end in users:
            end = work_end or WORK_END_TIME
            try:
                end_dt = datetime.strptime(end, "%H:%M")
            except ValueError:
                continue
            for offset in (15, 10, 5):
                target_hm = (end_dt - timedelta(minutes=offset)).strftime("%H:%M")
                if target_hm == now_hm:
                    cursor.execute(
                        "SELECT id FROM attendance WHERE user_id = ? AND date = ? "
                        "AND check_in_time IS NOT NULL AND check_out_time IS NULL",
                        (user_id, today_str)
                    )
                    if cursor.fetchone():
                        targets.append((user_id, full_name, offset))
                    break
        return targets


async def send_checkout_reminders():
    """Har daqiqada ishga tushadi: har bir xodimning SHAXSIY ish tugash
    vaqtiga nisbatan 15, 10 va 5 daqiqa qolganda (masalan 18:00 da tugasa -
    17:45, 17:50, 17:55 da) hali chiqmagan bo'lsa eslatma yuboradi."""
    now = get_now()
    today_str = now.strftime("%Y-%m-%d")
    now_hm = now.strftime("%H:%M")
    targets = await run_db(_get_checkout_reminder_targets_sync, today_str, now_hm)

    for user_id, full_name, offset in targets:
        try:
            await bot.send_message(
                chat_id=user_id,
                text=(
                    "Mana ish vaqti ham yakunlanmoqda, ketishda \"Ishdan ketdim\" "
                    "tugmasini bosib ketganligingizni tasdiqlash esdan chiqmasin"
                ),
            )
        except TelegramForbiddenError:
            logger.info(f"Foydalanuvchi {user_id} botni bloklagan, eslatma yuborilmadi.")
        except Exception as e:
            logger.error(f"Chiqish eslatmasini yuborishda xatolik ({user_id}): {e}")


async def send_daily_report():
    """Har kuni soat 21:00 da o'sha kunning to'liq davomat
    statistikasini Excel (.xlsx) formatida barcha adminlarga yuboradi."""
    today_str = get_now().strftime("%Y-%m-%d")
    file_path = f"Kunlik_Hisobot_{today_str}.xlsx"

    try:
        await generate_daily_excel_report(today_str, file_path)
    except Exception as e:
        logger.error(f"Kunlik hisobot generatsiyasida xatolik: {e}")
        return

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(
                chat_id=admin_id,
                document=FSInputFile(file_path),
                caption=f"📊 <b>{esc(today_str)}</b> kuni uchun yakuniy davomat statistikasi.",
                parse_mode="HTML"
            )
        except Exception as e:
            logger.error(f"Kunlik hisobotni yuborishda xatolik ({admin_id}): {e}")


async def send_admin_daily_summary():
    """Ish kuni yakunlanganda (18:05) adminlarga qisqa matnli xulosa yuboradi:
    necha kishi kelgan, necha kishi kelmagan, kimlar kech qolgan."""
    today_str = get_now().strftime("%Y-%m-%d")
    ishda, ketgan, kelmagan = await run_db(_get_employee_status_sync, today_str)
    kelganlar = ishda + [(name, ci, None) for name, ci, co in ketgan]
    total = len(kelganlar) + len(kelmagan)

    late_list = [(name, lateness) for name, _, lateness in ishda if lateness and lateness > 0]

    lines = [
        f"📋 <b>Kunlik xulosa</b> ({esc(today_str)}):\n",
        f"👥 Jami xodimlar: <b>{total}</b>",
        f"✅ Kelganlar: <b>{len(kelganlar)}</b>",
        f"⏳ Kelmaganlar: <b>{len(kelmagan)}</b>",
        f"⚠️ Kech qolganlar: <b>{len(late_list)}</b>",
    ]

    if kelmagan:
        lines.append("\n⏳ <b>Kelmaganlar:</b>")
        for _uid, name in kelmagan:
            lines.append(f"  • {esc(name)}")

    if late_list:
        lines.append("\n⚠️ <b>Kech qolganlar:</b>")
        for name, lateness in late_list:
            lines.append(f"  • {esc(name)} — {lateness} daqiqa")

    text = "\n".join(lines)
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=admin_id, text=text, parse_mode="HTML")
        except Exception as e:
            logger.error(f"Kunlik xulosani yuborishda xatolik ({admin_id}): {e}")


def _get_weekly_summary_sync(start_str: str, end_str: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_id, full_name FROM users WHERE is_approved = 1 ORDER BY full_name COLLATE NOCASE"
        )
        users = cursor.fetchall()

        results = []
        for user_id, full_name in users:
            if user_id in ADMIN_IDS:
                continue
            cursor.execute(
                "SELECT COUNT(id), COALESCE(SUM(lateness_minutes), 0), COALESCE(SUM(work_hours), 0.0) "
                "FROM attendance WHERE user_id = ? AND date >= ? AND date <= ?",
                (user_id, start_str, end_str)
            )
            days_came, total_lateness, total_hours = cursor.fetchone()
            results.append((full_name, days_came, total_lateness, total_hours))
        return results


async def send_admin_weekly_summary():
    """Har yakshanba kechqurun o'tgan hafta bo'yicha xodimlar kesimida qisqa xulosa yuboradi."""
    now = get_now()
    week_start = (now - timedelta(days=6)).strftime("%Y-%m-%d")
    week_end = now.strftime("%Y-%m-%d")

    results = await run_db(_get_weekly_summary_sync, week_start, week_end)

    lines = [f"📅 <b>Haftalik xulosa</b> ({esc(week_start)} — {esc(week_end)}):\n"]
    if not results:
        lines.append("<i>Xodimlar topilmadi.</i>")
    else:
        for full_name, days_came, total_lateness, total_hours in results:
            lines.append(
                f"👤 <b>{esc(full_name)}</b>: {days_came} kun kelgan, "
                f"{round(total_hours, 1)} soat ishlagan, {total_lateness} daq. kechikkan"
            )

    text = "\n".join(lines)
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=admin_id, text=text, parse_mode="HTML")
        except Exception as e:
            logger.error(f"Haftalik xulosani yuborishda xatolik ({admin_id}): {e}")


async def send_weekly_db_backup():
    """Qo'shimcha xavfsizlik chorasi sifatida, haftada bir marta (yakshanba)
    xom .db faylining to'liq nusxasi adminlarga yuboriladi."""
    now_str = get_now().strftime('%Y-%m-%d %H:%M')
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(
                chat_id=admin_id,
                document=FSInputFile(DB_PATH),
                caption=f"💾 <b>Haftalik zaxira nusxa (.db):</b> {esc(now_str)}",
                parse_mode="HTML"
            )
        except Exception as e:
            logger.error(f"Haftalik backup yuborishda xatolik ({admin_id}): {e}")


async def send_monthly_report_job():
    """Har oyning OXIRGI kuni soat 21:00 da shu oy uchun to'liq oylik hisobotni yuboradi."""
    now = get_now()
    month_prefix = now.strftime("%Y-%m")
    month_name = now.strftime("%B %Y")
    file_path = f"Oylik_Hisobot_{month_prefix}.xlsx"

    try:
        await generate_excel_report(month_prefix, file_path)
    except Exception as e:
        logger.error(f"Oylik hisobot generatsiyasida xatolik: {e}")
        return

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_document(
                chat_id=admin_id,
                document=FSInputFile(file_path),
                caption=f"📅 <b>{esc(month_name)}</b> oyi uchun avtomatik oylik va avans hisoboti.",
                parse_mode="HTML"
            )
        except Exception as e:
            logger.error(f"Avto hisobot yuborishda xatolik ({admin_id}): {e}")


# ================= /cancel BUYRUG'I =================
@dp.message(Command("cancel"))
async def cmd_cancel(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        await message.answer("ℹ️ Hozircha bekor qilinadigan amal yo'q.")
        return
    await state.clear()
    await message.answer(
        "❌ Amal bekor qilindi.",
        reply_markup=get_kb_for(message.from_user.id)
    )


# ================= START & USER APPROVAL =================
def _get_user_sync(user_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT is_approved FROM users WHERE user_id = ?", (user_id,))
        return cursor.fetchone()


def _insert_new_user_sync(user_id: int, full_name: str, is_approved: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO users (user_id, full_name, is_approved, monthly_salary, norm_days) VALUES (?, ?, ?, 0, 26)',
            (user_id, full_name, is_approved)
        )
        conn.commit()


@dp.message(CommandStart())
async def cmd_start(message: types.Message):
    user_id = message.from_user.id
    full_name = message.from_user.full_name

    user = await run_db(_get_user_sync, user_id)

    if not user:
        if user_id in ADMIN_IDS:
            await run_db(_insert_new_user_sync, user_id, full_name, 1)
            await message.answer(
                f"Assalomu alaykum, Admin <b>{esc(full_name)}</b>!",
                reply_markup=get_kb_for(user_id),
                parse_mode="HTML"
            )
            return

        await run_db(_insert_new_user_sync, user_id, full_name, 0)
        await message.answer(
            "⏳ <b>Arizangiz adminga yuborildi.</b> Admin tasdiqlashini kuting...",
            parse_mode="HTML"
        )

        approve_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=f"approve_{user_id}"),
             InlineKeyboardButton(text="❌ Rad etish", callback_data=f"reject_{user_id}")]
        ])
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(
                    admin_id,
                    f"🔔 <b>Yangi foydalanuvchi:</b>\n👤 F.I.Sh: {esc(full_name)}\n🆔 ID: <code>{user_id}</code>",
                    reply_markup=approve_kb,
                    parse_mode="HTML"
                )
            except Exception as e:
                logger.error(f"Admin({admin_id})ga xabar yuborishda xatolik: {e}")
        return

    if user[0] == 0:
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    await message.answer(
        f"Assalomu alaykum, <b>{esc(full_name)}</b>!",
        reply_markup=get_kb_for(user_id),
        parse_mode="HTML"
    )


def _approve_user_sync(user_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET is_approved = 1 WHERE user_id = ?", (user_id,))
        conn.commit()


@dp.callback_query(F.data.startswith("approve_"))
async def approve_user(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return

    user_id = int(callback.data.split("_")[1])
    await run_db(_approve_user_sync, user_id)

    await callback.message.edit_text(f"✅ Foydalanuvchi ({user_id}) tasdiqlandi!")
    try:
        await bot.send_message(
            user_id,
            "🎉 <b>Arizangiz tasdiqlandi!</b> Endi botdan foydalanishingiz mumkin.",
            reply_markup=user_keyboard,
            parse_mode="HTML"
        )
    except Exception as e:
        logger.error(f"Foydalanuvchi({user_id})ga tasdiqlash xabarini yuborishda xatolik: {e}")
    await callback.answer()


def _reject_user_sync(user_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
        conn.commit()


@dp.callback_query(F.data.startswith("reject_"))
async def reject_user(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return

    user_id = int(callback.data.split("_")[1])
    await run_db(_reject_user_sync, user_id)

    await callback.message.edit_text(f"❌ Foydalanuvchi ({user_id}) rad etildi va o'chirildi.")
    try:
        await bot.send_message(user_id, "❌ Afsuski, arizangiz admin tomonidan rad etildi.")
    except Exception as e:
        logger.error(f"Foydalanuvchi({user_id})ga rad javobini yuborishda xatolik: {e}")
    await callback.answer()


# ================= CHECK-IN & REASON FLOW =================
def _get_today_attendance_sync(user_id: int, today_str: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, check_in_time, check_out_time, break_start, break_end, break_minutes, "
            "checkin_source, checkout_source "
            "FROM attendance WHERE user_id = ? AND date = ?",
            (user_id, today_str)
        )
        return cursor.fetchone()


@dp.message(F.text == "🟢 Ishga keldim")
async def ask_location_for_checkin(message: types.Message):
    if not await is_approved(message.from_user.id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    today_str = get_now().strftime("%Y-%m-%d")
    record = await run_db(_get_today_attendance_sync, message.from_user.id, today_str)

    if record:
        checkin_source = record[6] if len(record) > 6 else "self"
        if checkin_source == "admin":
            await message.answer(
                f"ℹ️ Admin tomonidan siz bugun soat <b>{esc(record[1])}</b> da kelgan deb belgilangansiz.",
                parse_mode="HTML"
            )
        else:
            await message.answer(f"⚠️ Siz bugun soat <b>{esc(record[1])}</b> da kelgansiz!", parse_mode="HTML")
        return

    await message.answer("📍 Joylashuvni yuborish uchun pastdagi tugmani bosing:", reply_markup=location_keyboard)


def _checkout_sync(user_id: int, today_str: str, current_time_str: str):
    """Ishdan chiqishni yozib qo'yadi, agar tanaffus yopilmagan bo'lsa avtomatik yopadi.
    Natija: (status, payload)
      status == "no_checkin" -> yozuv topilmadi
      status == "ok" -> (net_work_hours, break_autoclosed(bool), break_minutes_added)
    """
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, check_in_time, break_minutes, break_start, break_end "
            "FROM attendance WHERE user_id = ? AND date = ?",
            (user_id, today_str)
        )
        record = cursor.fetchone()

        if not record:
            return "no_checkin", None

        record_id, check_in_time, break_minutes, break_start, break_end = record
        break_minutes = break_minutes or 0
        break_autoclosed = False
        added_break_minutes = 0

        now = get_now()

        # Agar tanaffus boshlangan-u, hali yopilmagan bo'lsa - avtomatik yopamiz
        if break_start and not break_end:
            break_start_dt = datetime.strptime(
                f"{today_str} {break_start}", "%Y-%m-%d %H:%M:%S"
            ).replace(tzinfo=TASHKENT_TZ)
            added_break_minutes = int((now - break_start_dt).total_seconds() / 60)
            break_minutes += added_break_minutes
            break_autoclosed = True
            cursor.execute(
                "UPDATE attendance SET break_end = ?, break_minutes = ? WHERE id = ?",
                (now.strftime("%H:%M:%S"), break_minutes, record_id)
            )

        check_in_dt = datetime.strptime(
            f"{today_str} {check_in_time}", "%Y-%m-%d %H:%M:%S"
        ).replace(tzinfo=TASHKENT_TZ)
        total_seconds = (now - check_in_dt).total_seconds()

        break_seconds = break_minutes * 60
        net_work_hours = round(max(0, total_seconds - break_seconds) / 3600, 2)

        cursor.execute(
            "UPDATE attendance SET check_out_time = ?, work_hours = ? WHERE id = ?",
            (current_time_str, net_work_hours, record_id)
        )
        conn.commit()

        cursor.execute("SELECT monthly_salary, norm_days FROM users WHERE user_id = ?", (user_id,))
        u_data = cursor.fetchone()
        if u_data:
            computed_norm = _norm_days_for_month(cursor, user_id, today_str[:7], u_data[1])
            u_data = (u_data[0], computed_norm)

    return "ok", (net_work_hours, break_autoclosed, added_break_minutes, u_data)


@dp.message(F.location, AttendanceState.waiting_for_checkout_location)
async def process_checkout_location(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        await state.clear()
        return

    user_lat, user_lon = message.location.latitude, message.location.longitude

    distance = calculate_distance(OFFICE_LAT, OFFICE_LON, user_lat, user_lon)
    kb = get_kb_for(user_id)

    if distance > ALLOWED_RADIUS_METERS:
        await message.answer(
            f"❌ <b>Joylashuv xatosi!</b> Ofisgacha masofa: <b>{int(distance)} metr</b>. "
            f"Ofis hududida bo'lishingiz kerak.",
            reply_markup=kb, parse_mode="HTML"
        )
        await state.clear()
        return

    now = get_now()
    today_str = now.strftime("%Y-%m-%d")
    current_time_str = now.strftime("%H:%M:%S")

    status, payload = await run_db(_checkout_sync, user_id, today_str, current_time_str)

    if status == "no_checkin":
        await message.answer("⚠️ Ishga kelganingiz haqida yozuv topilmadi.", reply_markup=kb)
        await state.clear()
        return

    net_work_hours, break_autoclosed, added_break_minutes, u_data = payload

    m_salary = u_data[0] if u_data else 0.0
    norm_d = u_data[1] if u_data and u_data[1] > 0 else 26

    hourly_rate = m_salary / (norm_d * 8) if (norm_d * 8) > 0 else 0
    daily_earned = net_work_hours * hourly_rate

    extra_note = ""
    if break_autoclosed:
        extra_note = (
            f"\nℹ️ Tanaffusdan qaytganingizni belgilamagan edingiz, shuning uchun "
            f"tanaffus avtomatik yopildi ({added_break_minutes} daqiqa qo'shildi)."
        )

    await message.answer(
        f"🔴 <b>Ishdan ketganingiz belgilandi!</b>\n"
        f"⏰ Ketish vaqti: <b>{current_time_str}</b>\n"
        f"⏱ Ishlangan net-vaqt: <b>{net_work_hours} soat</b>\n"
        f"💰 Bugungi ishlagan pulingiz: <b>{daily_earned:,.0f} so'm</b>"
        f"{extra_note}",
        reply_markup=kb,
        parse_mode="HTML"
    )
    await state.clear()


def _checkin_sync(user_id: int, today_str: str, current_time_str: str, lateness: int):
    """Check-in yozuvini kiritadi. UNIQUE indeks tufayli takroriy urinish IntegrityError beradi."""
    with get_db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute(
                "INSERT INTO attendance (user_id, date, check_in_time, lateness_minutes) VALUES (?, ?, ?, ?)",
                (user_id, today_str, current_time_str, lateness)
            )
        except sqlite3.IntegrityError:
            conn.rollback()
            return None
        attendance_id = cursor.lastrowid
        conn.commit()
        return attendance_id


def _get_user_work_hours_sync(user_id: int):
    """Xodimning shaxsiy ish boshlash/tugash vaqtini qaytaradi.
    Agar admin alohida belgilamagan bo'lsa, umumiy (global) sozlamaga qaytadi."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT work_start_time, work_end_time FROM users WHERE user_id = ?", (user_id,))
        row = cursor.fetchone()
        start = (row[0] if row and row[0] else WORK_START_TIME)
        end = (row[1] if row and row[1] else WORK_END_TIME)
        return start, end


@dp.message(F.location)
async def handle_location(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    user_lat, user_lon = message.location.latitude, message.location.longitude

    distance = calculate_distance(OFFICE_LAT, OFFICE_LON, user_lat, user_lon)
    kb = get_kb_for(user_id)

    if distance > ALLOWED_RADIUS_METERS:
        await message.answer(
            f"❌ <b>Joylashuv xatosi!</b> Ofisgacha masofa: <b>{int(distance)} metr</b>.",
            reply_markup=kb, parse_mode="HTML"
        )
        return

    now = get_now()
    today_str = now.strftime("%Y-%m-%d")
    current_time_str = now.strftime("%H:%M:%S")

    user_work_start, _ = await run_db(_get_user_work_hours_sync, user_id)
    work_start = datetime.strptime(f"{today_str} {user_work_start}", "%Y-%m-%d %H:%M").replace(tzinfo=TASHKENT_TZ)
    lateness = 0
    if now > work_start:
        lateness = int((now - work_start).total_seconds() / 60)

    attendance_id = await run_db(_checkin_sync, user_id, today_str, current_time_str, lateness)

    if attendance_id is None:
        # Allaqachon check-in qilingan (race condition oldini olindi)
        record = await run_db(_get_today_attendance_sync, user_id, today_str)
        check_in_time = record[1] if record else "?"
        checkin_source = record[6] if record and len(record) > 6 else "self"
        if checkin_source == "admin":
            text = f"ℹ️ Admin tomonidan siz bugun soat <b>{esc(check_in_time)}</b> da kelgan deb belgilangansiz."
        else:
            text = f"⚠️ Siz bugun soat <b>{esc(check_in_time)}</b> da kelgansiz!"
        await message.answer(text, reply_markup=kb, parse_mode="HTML")
        return

    if lateness > 0:
        await state.update_data(attendance_id=attendance_id, lateness=lateness, current_time=current_time_str)
        await message.answer(
            f"⚠️ <b>Siz {lateness} daqiqa kechikdingiz!</b>\n\nIltimos, kechikish sababini yozib qoldiring:",
            reply_markup=types.ReplyKeyboardRemove(),
            parse_mode="HTML"
        )
        await state.set_state(AttendanceState.waiting_for_reason)
    else:
        await message.answer(
            f"✅ <b>Ishga kelganingiz belgilandi!</b>\n⏰ Vaqt: <b>{current_time_str}</b>",
            reply_markup=kb, parse_mode="HTML"
        )


def _save_lateness_reason_sync(attendance_id: int, reason: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE attendance SET lateness_reason = ? WHERE id = ?", (reason, attendance_id))
        conn.commit()


@dp.message(AttendanceState.waiting_for_reason)
async def handle_lateness_reason(message: types.Message, state: FSMContext):
    reason = message.text
    data = await state.get_data()
    user_id = message.from_user.id

    await run_db(_save_lateness_reason_sync, data['attendance_id'], reason)

    kb = get_kb_for(user_id)
    await message.answer("✅ <b>Rahmat! Sabab saqlandi.</b>", reply_markup=kb, parse_mode="HTML")

    admin_alert = (
        f"🚨 <b>KECHIKISH VA SABAB!</b>\n\n"
        f"👤 Xodim: <b>{esc(message.from_user.full_name)}</b>\n"
        f"⏰ Kelgan vaqti: <b>{esc(data['current_time'])}</b>\n"
        f"⏱ Kechikish: <b>{data['lateness']} daqiqa</b>\n"
        f"📝 Sababi: <i>{esc(reason)}</i>"
    )
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(chat_id=admin_id, text=admin_alert, parse_mode="HTML")
        except Exception as e:
            logger.error(f"Admin({admin_id})ga kechikish xabarini yuborishda xatolik: {e}")

    await state.clear()


# ================= BREAK (TANAFFUS) HANDLERS =================
def _break_start_sync(user_id: int, today_str: str, now_str: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, check_in_time, break_start, check_out_time FROM attendance WHERE user_id = ? AND date = ?",
            (user_id, today_str)
        )
        record = cursor.fetchone()

        if not record:
            return "no_checkin", None
        if record[3] is not None:
            return "already_checked_out", None
        if record[2] is not None:
            return "already_on_break", record[2]

        cursor.execute("UPDATE attendance SET break_start = ? WHERE id = ?", (now_str, record[0]))
        conn.commit()
        return "ok", None


@dp.message(F.text == "☕ Tanaffusga chiqdim")
async def handle_break_start(message: types.Message):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    today_str = get_now().strftime("%Y-%m-%d")
    now_str = get_now().strftime("%H:%M:%S")

    status, extra = await run_db(_break_start_sync, user_id, today_str, now_str)

    if status == "no_checkin":
        await message.answer("⚠️ Avval ishga kelganingizni belgilang!")
        return
    if status == "already_checked_out":
        await message.answer("⚠️ Siz bugun allaqachon ishdan ketgansiz!")
        return
    if status == "already_on_break":
        await message.answer(f"⚠️ Siz allaqachon soat <b>{esc(extra)}</b> da tanaffusga chiqqansiz!", parse_mode="HTML")
        return

    await message.answer(
        f"☕ <b>Tanaffus vaqti boshlandi!</b>\nSoat: <b>{now_str}</b>\nYoqimli hordiq chiqaring!",
        parse_mode="HTML"
    )


def _break_end_sync(user_id: int, today_str: str, now_str: str, now_dt: datetime):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, break_start, break_end, break_minutes FROM attendance WHERE user_id = ? AND date = ?",
            (user_id, today_str)
        )
        record = cursor.fetchone()

        if not record or not record[1]:
            return "not_on_break", None
        if record[2] is not None:
            return "already_returned", None

        break_start_dt = datetime.strptime(
            f"{today_str} {record[1]}", "%Y-%m-%d %H:%M:%S"
        ).replace(tzinfo=TASHKENT_TZ)
        this_break_minutes = int((now_dt - break_start_dt).total_seconds() / 60)
        total_break_minutes = (record[3] or 0) + this_break_minutes

        cursor.execute(
            "UPDATE attendance SET break_end = ?, break_minutes = ? WHERE id = ?",
            (now_str, total_break_minutes, record[0])
        )
        conn.commit()
        return "ok", this_break_minutes


@dp.message(F.text == "🏢 Tanaffusdan qaytdim")
async def handle_break_end(message: types.Message):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    now = get_now()
    today_str = now.strftime("%Y-%m-%d")
    now_str = now.strftime("%H:%M:%S")

    status, this_break_minutes = await run_db(_break_end_sync, user_id, today_str, now_str, now)

    if status == "not_on_break":
        await message.answer("⚠️ Siz tanaffusga chiqqaningizni belgilamagansiz!")
        return
    if status == "already_returned":
        await message.answer("⚠️ Siz tanaffusdan qaytganingizni belgilab bo'lgansiz!")
        return

    await message.answer(
        f"🏢 <b>Ishga qaytganingiz belgilandi!</b>\nTanaffus davomiyligi: <b>{this_break_minutes} daqiqa</b>.",
        parse_mode="HTML"
    )


# ================= CHECK-OUT HANDLER =================
@dp.message(F.text == "🔴 Ishdan ketdim")
async def handle_checkout_request(message: types.Message, state: FSMContext):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    today_str = get_now().strftime("%Y-%m-%d")

    record = await run_db(_get_today_attendance_sync, user_id, today_str)

    if not record:
        await message.answer("⚠️ Siz bugun ishga kelganingizni belgilamagansiz!")
        return
    elif record[2] is not None:
        checkout_source = record[7] if len(record) > 7 else "self"
        if checkout_source == "admin":
            await message.answer(
                f"ℹ️ Admin tomonidan siz bugun soat <b>{esc(record[2])}</b> da ketgan deb belgilangansiz.",
                parse_mode="HTML"
            )
        else:
            await message.answer(f"⚠️ Siz bugun soat <b>{esc(record[2])}</b> da ketganingiz belgilangan!", parse_mode="HTML")
        return

    await state.set_state(AttendanceState.waiting_for_checkout_location)
    await message.answer("📍 Ishdan ketishni tasdiqlash uchun joylashuvni yuboring:", reply_markup=location_keyboard)


# ================= STATISTIKA & OYLIK =================
def _get_user_stats_sync(user_id: int, month_prefix: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COUNT(id), COALESCE(SUM(lateness_minutes), 0), COALESCE(SUM(work_hours), 0.0), COALESCE(SUM(break_minutes), 0)
            FROM attendance WHERE user_id = ? AND date LIKE ?
        ''', (user_id, f"{month_prefix}%"))
        row = cursor.fetchone()

        cursor.execute("SELECT monthly_salary, norm_days FROM users WHERE user_id = ?", (user_id,))
        salary_row = cursor.fetchone()
        if salary_row:
            computed_norm = _norm_days_for_month(cursor, user_id, month_prefix, salary_row[1])
            salary_row = (salary_row[0], computed_norm)

        return row, salary_row


@dp.message(F.text == "📊 Mening statistikam")
async def cmd_user_stats(message: types.Message):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    current_month_prefix = get_now().strftime("%Y-%m")

    row, salary_row = await run_db(_get_user_stats_sync, user_id, current_month_prefix)

    monthly_salary = salary_row[0] if salary_row else 0.0
    norm_days = salary_row[1] if salary_row and salary_row[1] > 0 else 26

    hourly_rate = monthly_salary / (norm_days * 8) if (norm_days * 8) > 0 else 0.0
    total_work_hours = row[2]
    total_earned = total_work_hours * hourly_rate

    stats_text = (
        f"📊 <b>Shu oy bo'yicha statistikangiz:</b>\n\n"
        f"📅 Ishga kelgan kunlar: <b>{row[0]} kun</b>\n"
        f"⏱ Jami ishlangan net-soat: <b>{round(total_work_hours, 1)} soat</b>\n"
        f"☕ Jami tanaffus vaqti: <b>{row[3]} daqiqa</b>\n"
        f"⚠️ Jami kechikish: <b>{row[1]} daqiqa</b>\n"
        f"💵 Belgilangan oylik: <b>{monthly_salary:,.0f} so'm</b> (Norma: {norm_days} kun)\n"
        f"💰 <b>Topilgan pul (soatbay): {total_earned:,.0f} so'm</b>"
    )
    await message.answer(stats_text, parse_mode="HTML")


def _get_salary_details_sync(user_id: int, month_prefix: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT full_name, monthly_salary, norm_days FROM users WHERE user_id = ?", (user_id,))
        user_info = cursor.fetchone()

        if not user_info:
            return None

        computed_norm = _norm_days_for_month(cursor, user_id, month_prefix, user_info[2])
        user_info = (user_info[0], user_info[1], computed_norm)

        cursor.execute(
            "SELECT COALESCE(SUM(work_hours), 0.0) FROM attendance WHERE user_id = ? AND date LIKE ?",
            (user_id, f"{month_prefix}%")
        )
        total_hours = cursor.fetchone()[0]

        cursor.execute(
            "SELECT COALESCE(SUM(amount), 0.0) FROM advances WHERE user_id = ? AND date LIKE ?",
            (user_id, f"{month_prefix}%")
        )
        total_advance = cursor.fetchone()[0]

        return user_info, total_hours, total_advance


@dp.message(F.text == "💰 Mening oyligim")
async def cmd_my_salary(message: types.Message):
    user_id = message.from_user.id

    if not await is_approved(user_id):
        await message.answer("⏳ Sizning arizangiz hali admin tomonidan tasdiqlanmagan.")
        return

    current_month_prefix = get_now().strftime("%Y-%m")

    result = await run_db(_get_salary_details_sync, user_id, current_month_prefix)

    if not result:
        await message.answer("⚠️ Ma'lumot topilmadi.")
        return

    user_info, total_hours, total_advance = result
    full_name, m_salary, norm_days = user_info

    hourly_rate = m_salary / (norm_days * 8) if (norm_days * 8) > 0 else 0.0
    gross_earned = total_hours * hourly_rate
    net_payable = gross_earned - total_advance

    text = (
        f"💰 <b>Sizning joriy oylik hisob-kitobingiz:</b>\n\n"
        f"👤 Xodim: <b>{esc(full_name)}</b>\n"
        f"💵 Belgilangan oylik: <b>{m_salary:,.0f} so'm</b>\n"
        f"📅 Ish kuni normasi: <b>{norm_days} kun</b>\n"
        f"⏱ 1 soatlik ish haqi: <b>{hourly_rate:,.0f} so'm</b>\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"📈 Shu oygacha ishlangan soat: <b>{round(total_hours, 1)} soat</b>\n"
        f"💵 Ishlangan umumiy pul: <b>{gross_earned:,.0f} so'm</b>\n"
        f"💸 Olingan avanslar: <b>{total_advance:,.0f} so'm</b>\n"
        f"━━━━━━━━━━━━━━━━━━━\n"
        f"💼 <b>Qo'lga tegadigan sof oylik: {net_payable:,.0f} so'm</b>"
    )
    await message.answer(text, parse_mode="HTML")


@dp.message(F.text == "📈 Barcha davomat hisoboti (Excel)")
async def cmd_admin_report(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    month_prefix = get_now().strftime("%Y-%m")
    file_path = f"Oylik_Hisobot_{get_now().strftime('%Y%m%d_%H%M')}.xlsx"

    try:
        await generate_excel_report(month_prefix, file_path)
        await message.answer_document(FSInputFile(file_path), caption="📊 Xodimlarning oylik, avans va sof maosh hisoboti")
    except Exception as e:
        logger.error(f"Excel hisobot yaratishda xatolik: {e}")
        await message.answer("❌ Hisobot yaratishda xatolik yuz berdi. Qaytadan urinib ko'ring.")


# ================= XABARNOMA (BROADCAST) =================
def _get_approved_users_sync():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, full_name FROM users WHERE is_approved = 1")
        return cursor.fetchall()


def build_checkin_reminder_text(full_name: str) -> str:
    return (
        f"Xayrli kun, <b>{esc(full_name)}</b>! ☀️\n"
        f"Ofisga kelgach, <b>🟢 Ishga keldim</b> tugmasini bosishni unutmang!"
    )


def build_checkout_reminder_text(full_name: str) -> str:
    return (
        f"Hurmatli <b>{esc(full_name)}</b>,\n"
        f"Ish vaqti tugagach, <b>🔴 Ishdan ketdim</b> tugmasini bosishni unutmang!"
    )


def build_custom_broadcast_text(full_name: str, custom_text: str) -> str:
    return f"📢 <b>E'lon:</b>\n\n{esc(custom_text)}"


async def broadcast_message(text_builder) -> tuple[int, int]:
    """text_builder(full_name) -> HTML matn. Har bir tasdiqlangan foydalanuvchiga
    shaxsiylashtirilgan xabar yuboradi. (muvaffaqiyatli, muvaffaqiyatsiz) sonini qaytaradi."""
    users = await run_db(_get_approved_users_sync)
    success, failed = 0, 0
    for user_id, full_name in users:
        try:
            await bot.send_message(
                chat_id=user_id,
                text=text_builder(full_name),
                reply_markup=get_kb_for(user_id),
                parse_mode="HTML"
            )
            success += 1
        except TelegramForbiddenError:
            failed += 1
            logger.info(f"Foydalanuvchi {user_id} botni bloklagan, xabarnoma yuborilmadi.")
        except Exception as e:
            failed += 1
            logger.error(f"Xabarnoma yuborishda xatolik ({user_id}): {e}")
    return success, failed


@dp.message(F.text == "📢 Xabarnoma yuborish")
async def broadcast_menu(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="☀️ Standart: Kelish eslatmasi", callback_data="broadcast_checkin")],
        [InlineKeyboardButton(text="🌙 Standart: Ketish eslatmasi", callback_data="broadcast_checkout")],
        [InlineKeyboardButton(text="✍️ Maxsus xabar yozish", callback_data="broadcast_custom")]
    ])
    await message.answer(
        "📢 <b>Xabarnoma yuborish:</b>\n\n"
        "Standart eslatmalardan birini tanlang yoki o'zingiz xabar yozing.\n"
        "Xabar barcha tasdiqlangan xodimlarga yuboriladi.",
        reply_markup=kb,
        parse_mode="HTML"
    )


@dp.callback_query(F.data == "broadcast_checkin")
async def broadcast_checkin_callback(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return
    await callback.answer("Yuborilmoqda...")
    success, failed = await broadcast_message(build_checkin_reminder_text)
    await callback.message.answer(
        f"✅ Kelish eslatmasi <b>{success}</b> ta xodimga yuborildi."
        + (f"\n⚠️ <b>{failed}</b> ta xodimga yetkazib bo'lmadi." if failed else ""),
        parse_mode="HTML"
    )


@dp.callback_query(F.data == "broadcast_checkout")
async def broadcast_checkout_callback(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return
    await callback.answer("Yuborilmoqda...")
    success, failed = await broadcast_message(build_checkout_reminder_text)
    await callback.message.answer(
        f"✅ Ketish eslatmasi <b>{success}</b> ta xodimga yuborildi."
        + (f"\n⚠️ <b>{failed}</b> ta xodimga yetkazib bo'lmadi." if failed else ""),
        parse_mode="HTML"
    )


@dp.callback_query(F.data == "broadcast_custom")
async def broadcast_custom_start(callback: types.CallbackQuery, state: FSMContext):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return
    await callback.message.answer(
        "✍️ Iltimos, barcha xodimlarga yuboriladigan xabar matnini kiriting:\n"
        "(Bekor qilish uchun /cancel)"
    )
    await state.set_state(AttendanceState.waiting_for_broadcast_text)
    await callback.answer()


@dp.message(AttendanceState.waiting_for_broadcast_text)
async def process_broadcast_custom_text(message: types.Message, state: FSMContext):
    custom_text = message.text.strip()
    if not custom_text:
        await message.answer("❌ Xabar matni bo'sh bo'lishi mumkin emas. Qaytadan kiriting:")
        return

    await state.clear()
    await message.answer("⏳ Xabar yuborilmoqda...")

    success, failed = await broadcast_message(lambda name: build_custom_broadcast_text(name, custom_text))
    await message.answer(
        f"✅ Xabar <b>{success}</b> ta xodimga yuborildi."
        + (f"\n⚠️ <b>{failed}</b> ta xodimga yetkazib bo'lmadi." if failed else ""),
        parse_mode="HTML"
    )


# ================= XODIMLAR (REAL VAQTDAGI HOLAT) =================
def _get_employee_status_sync(today_str: str):
    """Har bir (admin bo'lmagan) tasdiqlangan xodim uchun bugungi holatni qaytaradi:
    'ishda' (kelgan, hali ketmagan), 'ketgan' (kelib-ketgan) yoki 'kelmagan'."""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_id, full_name, work_start_time FROM users "
            "WHERE is_approved = 1 ORDER BY full_name COLLATE NOCASE"
        )
        users = cursor.fetchall()

        ishda, ketgan, kelmagan = [], [], []
        for user_id, full_name, work_start in users:
            if user_id in ADMIN_IDS:
                continue
            cursor.execute(
                "SELECT check_in_time, check_out_time, lateness_minutes FROM attendance "
                "WHERE user_id = ? AND date = ?",
                (user_id, today_str)
            )
            record = cursor.fetchone()
            if not record:
                kelmagan.append((user_id, full_name))
            elif record[1] is None:
                ishda.append((full_name, record[0], record[2]))
            else:
                ketgan.append((full_name, record[0], record[1]))
        return ishda, ketgan, kelmagan


@dp.message(F.text == "👥 Xodimlar")
async def cmd_employee_status(message: types.Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    today_str = get_now().strftime("%Y-%m-%d")
    ishda, ketgan, kelmagan = await run_db(_get_employee_status_sync, today_str)

    now_str = get_now().strftime("%H:%M")
    lines = [f"👥 <b>Xodimlar holati</b> (hozir {now_str}):\n"]

    lines.append(f"🟢 <b>Ishda ({len(ishda)}):</b>")
    if ishda:
        for name, check_in, lateness in ishda:
            late_note = f" (⚠️ {lateness} daq. kech)" if lateness and lateness > 0 else ""
            lines.append(f"  • {esc(name)} — {esc(check_in)[:5]} dan{late_note}")
    else:
        lines.append("  <i>Yo'q</i>")

    lines.append(f"\n🔴 <b>Ketgan ({len(ketgan)}):</b>")
    if ketgan:
        for name, check_in, check_out in ketgan:
            lines.append(f"  • {esc(name)} — {esc(check_in)[:5]} - {esc(check_out)[:5]}")
    else:
        lines.append("  <i>Yo'q</i>")

    lines.append(f"\n⏳ <b>Kelmagan ({len(kelmagan)}):</b>")
    if kelmagan:
        for _uid, name in kelmagan:
            lines.append(f"  • {esc(name)}")
    else:
        lines.append("  <i>Yo'q</i>")

    await message.answer("\n".join(lines), parse_mode="HTML")

    if kelmagan:
        nudge_buttons = [
            [InlineKeyboardButton(text=f"🔔 {name}", callback_data=f"nudge_{uid}")]
            for uid, name in kelmagan
        ]
        nudge_buttons.append(
            [InlineKeyboardButton(text="🔔 Hammasiga yuborish", callback_data="nudge_all")]
        )
        await message.answer(
            "Kelmagan xodimga \"ishga kelgan bo'lsangiz belgilashni unutmang\" "
            "eslatmasini yuborish uchun tanlang:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=nudge_buttons)
        )


@dp.callback_query(F.data.startswith("nudge_"))
async def nudge_employee(callback: types.CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Sizda ruxsat yo'q!", show_alert=True)
        return

    target = callback.data.split("_", 1)[1]
    text = "🔔 Ishga kelgan bo'lsangiz, belgilashni unutmang!"
    today_str = get_now().strftime("%Y-%m-%d")

    if target == "all":
        _, _, kelmagan = await run_db(_get_employee_status_sync, today_str)
        sent = 0
        for uid, _name in kelmagan:
            try:
                await bot.send_message(chat_id=uid, text=text)
                sent += 1
            except Exception as e:
                logger.error(f"Eslatma yuborishda xatolik ({uid}): {e}")
        await callback.answer(f"{sent} ta xodimga eslatma yuborildi.", show_alert=True)
    else:
        try:
            uid = int(target)
            await bot.send_message(chat_id=uid, text=text)
            await callback.answer("Eslatma yuborildi.", show_alert=True)
        except Exception as e:
            logger.error(f"Eslatma yuborishda xatolik ({target}): {e}")
            await callback.answer(
                "Xabar yuborilmadi (xodim botni bloklagan bo'lishi mumkin).",
                show_alert=True
            )


# ================= XODIMLARNI BOSHQARISH (endi to'liq Web Panel orqali) =================
# Eslatma: Botdagi "⚙️ Admin Panel" tugmasi va unga tegishli barcha
# Telegram-ichi boshqaruv oqimi (xodim qo'shish, ismini/oylik/ish vaqtini
# o'zgartirish, avans berish, davomatni qo'lda kiritish, o'chirish) olib
# tashlandi - shu funksiyalarning barchasi endi 🖥 Web Panel'dagi
# "⚙️ Xodimni boshqarish" bo'limi orqali bajariladi (pastdagi Flask API'lar
# aynan shu DB funksiyalarini chaqiradi). Yangi ro'yxatdan o'tish so'rovlarini
# tasdiqlash/rad etish esa avvalgidek Telegram orqali (✅/❌ tugmalari bilan)
# davom etadi - bu Admin Panel tugmasiga bog'liq bo'lmagan alohida oqim edi.
def _get_all_users_sync():
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT user_id, full_name, is_approved, monthly_salary "
            "FROM users ORDER BY full_name COLLATE NOCASE"
        )
        return cursor.fetchall()


def _rename_user_sync(user_id: int, new_name: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET full_name = ? WHERE user_id = ?", (new_name, user_id))
        conn.commit()


def _upsert_user_sync(user_id: int, name: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO users (user_id, full_name, is_approved, monthly_salary, norm_days) VALUES (?, ?, 1, 0.0, 26) "
            "ON CONFLICT(user_id) DO UPDATE SET full_name = excluded.full_name, is_approved = 1",
            (user_id, name)
        )
        conn.commit()


def _set_salary_sync(user_id: int, salary: float):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET monthly_salary = ? WHERE user_id = ?", (salary, user_id))
        conn.commit()


def _set_work_time_sync(user_id: int, start_time: str, end_time: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET work_start_time = ?, work_end_time = ? WHERE user_id = ?",
            (start_time, end_time, user_id)
        )
        conn.commit()


def _admin_set_attendance_sync(user_id: int, today_str: str, check_in_time: str, check_out_time: str | None, work_start: str):
    with get_db() as conn:
        cursor = conn.cursor()

        work_start_dt = datetime.strptime(f"{today_str} {work_start}", "%Y-%m-%d %H:%M")
        check_in_dt = datetime.strptime(f"{today_str} {check_in_time}", "%Y-%m-%d %H:%M")
        lateness = max(0, int((check_in_dt - work_start_dt).total_seconds() / 60))

        work_hours = 0.0
        if check_out_time:
            check_out_dt = datetime.strptime(f"{today_str} {check_out_time}", "%Y-%m-%d %H:%M")
            work_hours = round(max(0, (check_out_dt - check_in_dt).total_seconds()) / 3600, 2)

        cursor.execute("SELECT id FROM attendance WHERE user_id = ? AND date = ?", (user_id, today_str))
        existing = cursor.fetchone()

        check_in_str = f"{check_in_time}:00"
        check_out_str = f"{check_out_time}:00" if check_out_time else None

        if existing:
            if check_out_time:
                cursor.execute(
                    "UPDATE attendance SET check_in_time = ?, check_out_time = ?, lateness_minutes = ?, "
                    "work_hours = ?, checkin_source = 'admin', checkout_source = 'admin' WHERE id = ?",
                    (check_in_str, check_out_str, lateness, work_hours, existing[0])
                )
            else:
                cursor.execute(
                    "UPDATE attendance SET check_in_time = ?, lateness_minutes = ?, "
                    "checkin_source = 'admin' WHERE id = ?",
                    (check_in_str, lateness, existing[0])
                )
        else:
            cursor.execute(
                "INSERT INTO attendance (user_id, date, check_in_time, check_out_time, lateness_minutes, "
                "work_hours, checkin_source, checkout_source) VALUES (?, ?, ?, ?, ?, ?, 'admin', ?)",
                (user_id, today_str, check_in_str, check_out_str, lateness, work_hours,
                 'admin' if check_out_time else 'self')
            )
        conn.commit()


def _add_advance_sync(user_id: int, amount: float, date_str: str):
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO advances (user_id, amount, date) VALUES (?, ?, ?)", (user_id, amount, date_str))
        conn.commit()


def _delete_user_sync(user_id: int):
    with get_db() as conn:
        cursor = conn.cursor()
        # FOREIGN KEY cheklovi yoqilgani uchun avval bog'liq (child) yozuvlarni,
        # so'ngra users (parent) yozuvini o'chiramiz.
        cursor.execute("DELETE FROM attendance WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM advances WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM work_schedule WHERE user_id = ?", (user_id,))
        cursor.execute("DELETE FROM users WHERE user_id = ?", (user_id,))
        conn.commit()


# ================= MAIN FUNCTION =================
async def on_startup():
    logger.info("Bot ishga tushmoqda...")
    logger.info(f"Adminlar soni: {len(ADMIN_IDS)}")


async def main():
    global scheduler
    # Scheduler shu yerda, bot o'zining HAQIQIY (joriy) event loop'ini olgandan
    # keyin yaratiladi - shunday qilib add_job()/wakeup() har doim to'g'ri,
    # ochiq loop'ga ishlaydi (yuqoridagi izohga qarang).
    scheduler = AsyncIOScheduler(timezone="Asia/Tashkent", event_loop=asyncio.get_running_loop())

    await on_startup()

    # Ishga kelish/ketish eslatmalari endi har bir xodimning SHAXSIY ish
    # vaqtiga nisbatan hisoblanadi (masalan 09:00 dan ishlasa - 8:45/8:50/9:00 da,
    # 18:00 da tugasa - 17:45/17:50/17:55 da), shuning uchun har daqiqada tekshiriladi.
    scheduler.add_job(send_checkin_reminders, trigger="cron", day_of_week="mon-sat", minute="*")
    scheduler.add_job(send_checkout_reminders, trigger="cron", day_of_week="mon-sat", minute="*")
    scheduler.add_job(send_admin_daily_summary, trigger="cron", day_of_week="mon-sat", hour=18, minute=5)
    scheduler.add_job(send_admin_weekly_summary, trigger="cron", day_of_week="sun", hour=20, minute=0)
    scheduler.add_job(send_daily_report, trigger="cron", hour=21, minute=0)
    scheduler.add_job(send_weekly_db_backup, trigger="cron", day_of_week="sun", hour=23, minute=30)
    scheduler.add_job(send_monthly_report_job, trigger="cron", day="last", hour=21, minute=0)

    scheduler.start()
    try:
        # handle_signals=False SHART: bot bu yerda cPanel/Passenger'ning
        # asosiy (main) thread'ida emas, alohida background thread'da
        # ishlaydi (passenger_wsgi.py), aiogram esa SIGINT/SIGTERM signal
        # handler'larni faqat asosiy interpretator'ning asosiy thread'ida
        # o'rnata oladi - aks holda "set_wakeup_fd only works in main
        # thread of the main interpreter" xatosi bilan yiqiladi.
        await dp.start_polling(bot, handle_signals=False)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit):
        logger.info("Bot to'xtatildi.")